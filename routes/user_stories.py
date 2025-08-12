from flask import Blueprint, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from sqlalchemy import or_
from models import db, UserStory, User, SystemFeature, ProjectInfo, Sprint, SprintBacklog, UserRole, Role, \
    ProductBacklog
from utils import check_system_feature_access, check_user_role
from decorators import check_access_blueprint
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO

user_stories_bp = Blueprint('user_stories', __name__)

def generate_story_id(product_backlog_id):
    """
    生成用户故事编号
    规则：US_需求ID(3位)_序号(3位)
    """
    # 获取需求ID，保持3位长度，不够前面补零
    backlog_part = str(product_backlog_id).zfill(3)

    # 查询当前需求下最大的用户故事序号
    stories = UserStory.query.filter_by(product_backlog_id=product_backlog_id).all()

    if stories:
        # 查找最大的序号并加1
        max_seq = 0
        for story in stories:
            if story.story_id and story.story_id.startswith(f"US_{backlog_part}_"):
                try:
                    seq = int(story.story_id.split('_')[2])
                    if seq > max_seq:
                        max_seq = seq
                except (IndexError, ValueError):
                    # 如果解析失败，跳过该故事
                    pass
        next_seq = max_seq + 1
    else:
        # 如果当前需求没有用户故事，默认值为1
        next_seq = 1

    # 序号保持3位长度，不够前面补零
    seq_part = str(next_seq).zfill(3)

    # 拼接生成故事编号
    story_id = f"US_{backlog_part}_{seq_part}"
    return story_id


# 应用权限检查装饰器
@user_stories_bp.before_request
@check_access_blueprint('user_stories')
def check_access():
    pass  # 装饰器会处理权限检查逻辑

@user_stories_bp.route('/user_stories')
def user_stories():
    # 检查权限
    if not check_system_feature_access(session, 'user_stories.user_stories'):
        return redirect(url_for('auth.index'))
    
    # 获取所有项目（根节点）
    projects = ProjectInfo.query.filter_by(parent_id=None).order_by(ProjectInfo.order).all()
    
    # 获取所有用户故事
    user_stories = UserStory.query.order_by(UserStory.created_at.desc()).all()
    
    # 获取所有非管理员用户，用于分配负责人
    # 首先获取所有用户
    all_users = User.query.all()
    # 然后过滤掉具有admin角色的用户
    users = []
    for user in all_users:
        if not check_user_role(user.id, 'admin'):
            users.append(user)
    
    return render_template('user_stories.html', 
                          user_stories=user_stories, 
                          users=users,
                          projects=projects)

@user_stories_bp.route('/get_project_tree/<int:project_id>')
def get_project_tree(project_id):
    """获取指定项目的菜单和页面树形结构"""
    # 检查权限
    if not check_system_feature_access(session, 'user_stories.user_stories'):
        return jsonify({'success': False, 'message': '权限不足'})
    
    # 获取项目下的所有菜单和页面节点
    project = db.session.get(ProjectInfo, project_id)
    if not project:
        return jsonify({'success': False, 'message': '项目不存在'})
    
    nodes = ProjectInfo.query.filter(
        ProjectInfo.path.like(f"/{project.name}%")
    ).filter(
        ProjectInfo.node_type.in_(['menu', 'page'])
    ).order_by(ProjectInfo.path).all()
    
    # 构建树形结构
    def build_tree(parent_path=f"/{project.name}"):
        tree = []
        for node in nodes:
            # 检查是否为当前层级的直接子节点
            node_path_parts = node.path.split('/')
            parent_path_parts = parent_path.split('/')
            
            # 如果节点路径比父路径多一级，则为直接子节点
            if len(node_path_parts) == len(parent_path_parts) + 1 and node.path.startswith(parent_path + "/"):
                children = build_tree(node.path)
                tree.append({
                    'id': node.id,
                    'name': node.name,
                    'node_type': node.node_type,
                    'path': node.path,
                    'children': children if children else []
                })
        return tree
    
    # 构建项目下的树形结构
    project_tree = build_tree(f"/{project.name}")
    
    return jsonify({
        'success': True,
        'tree': project_tree
    })

@user_stories_bp.route('/get_product_backlogs')
def get_product_backlogs():
    """根据项目ID和状态获取产品待办事项列表"""
    # 检查权限
    if not check_system_feature_access(session, 'user_stories.user_stories'):
        return jsonify({'success': False, 'message': '权限不足'})

    # 获取查询参数
    project_id = request.args.get('project_id', type=int)
    status = request.args.get('status', type=str)

    # 构建查询
    query = ProductBacklog.query

    # 根据项目ID筛选
    if project_id:
        query = query.filter_by(project_id=project_id)

    # 根据状态筛选
    if status:
        query = query.filter_by(status=status)

    # 执行查询并按创建时间倒序排列
    product_backlogs = query.order_by(ProductBacklog.created_at.desc()).all()

    # 转换为字典格式
    backlogs_data = []
    for backlog in product_backlogs:
        backlogs_data.append({
            'id': backlog.id,
            'requirement_id': backlog.requirement_id,
            'title': backlog.title,
            'priority': backlog.priority,
            'status': backlog.status,
            'created_at': backlog.created_at.strftime('%Y-%m-%d %H:%M:%S') if backlog.created_at else ''
        })

    return jsonify({
        'success': True,
        'backlogs': backlogs_data
    })

@user_stories_bp.route('/get_user_stories_by_product_backlog/<int:product_backlog_id>')
def get_user_stories_by_product_backlog_id(product_backlog_id):
    """根据产品待办事项ID获取关联的用户故事"""
    # 检查权限
    if not check_system_feature_access(session, 'user_stories.user_stories'):
        return jsonify({'success': False, 'message': '权限不足'})

    # 获取产品待办事项
    product_backlog = db.session.get(ProductBacklog, product_backlog_id)
    if not product_backlog:
        return jsonify({'success': False, 'message': '产品待办事项不存在'})

    # 获取产品待办事项关联的用户故事
    user_stories = UserStory.query.filter_by(product_backlog_id=product_backlog_id).order_by(
        UserStory.created_at.desc()).all()

    # 获取所有非管理员用户，用于显示负责人信息
    users = []
    all_users = User.query.all()
    for user in all_users:
        if not check_user_role(user.id, 'admin'):
            users.append(user)
    
    # 转换用户故事为字典格式
    stories_data = []
    users_dict = {user.id: user for user in users}
    
    for story in user_stories:
        # 获取用户故事关联的迭代信息
        sprint_backlog = story.sprint_backlogs[0] if story.sprint_backlogs else None
        sprint_name = sprint_backlog.sprint.name if sprint_backlog else None
        
        stories_data.append({
            'id': story.id,
            'story_id': story.story_id or '',
            'title': story.title,
            'description': story.description or '',
            'acceptance_criteria': story.acceptance_criteria or '',
            'effort': story.effort or '',
            'priority': story.priority,
            'sprint': sprint_name or '',
            'created_at': story.created_at.strftime('%Y-%m-%d %H:%M:%S') if story.created_at else ''
        })
    
    # 转换用户为字典格式
    users_data = [{'id': user.id, 'name': user.name} for user in users]
    
    return jsonify({
        'success': True,
        'user_stories': stories_data,
        'users': users_data,
        'product_backlog_id': product_backlog_id
    })


@user_stories_bp.route('/add_user_story/<int:product_backlog_id>', methods=['GET', 'POST'])
def add_user_story(product_backlog_id):
    # 检查权限
    if not check_system_feature_access(session, 'user_stories.user_stories'):
        return jsonify({'success': False, 'message': '权限不足'})

    # 获取产品待办项
    product_backlog = db.session.get(ProductBacklog, product_backlog_id)
    if not product_backlog:
        return jsonify({'success': False, 'message': '产品待办事项不存在'})

    # 获取所有非管理员用户（通过角色关联判断）
    users = User.query.join(UserRole).filter(Role.name != 'admin').all()
    
    if request.method == 'POST':
        # 处理表单提交
        story_id = request.form.get('story_id', '').strip()
        title = request.form.get('title', '').strip()
        description = request.form.get('description', '').strip()
        acceptance_criteria = request.form.get('acceptance_criteria', '').strip()
        effort = request.form.get('effort', '').strip()
        priority = request.form.get('priority', 'P3')
        
        # 验证必填字段
        if not title:
            return jsonify({'success': False, 'message': '故事标题不能为空'})
        
        try:
            # 如果用户没有提供故事编号，则自动生成
            if not story_id:
                story_id = generate_story_id(product_backlog_id)
            
            # 创建新的用户故事
            user_story = UserStory(
                story_id=story_id,
                title=title,
                description=description or None,
                acceptance_criteria=acceptance_criteria or None,
                priority=priority,
                product_backlog_id=product_backlog_id
            )
            
            # 处理工作量
            if effort:
                try:
                    user_story.effort = float(effort)
                except ValueError:
                    return jsonify({'success': False, 'message': '工作量必须是有效的数字'})
            
            db.session.add(user_story)
            db.session.commit()
            
            # 返回新创建的用户故事信息
            story_data = {
                'id': user_story.id,
                'story_id': user_story.story_id or '',
                'title': user_story.title,
                'description': user_story.description or '',
                'acceptance_criteria': user_story.acceptance_criteria or '',
                'effort': user_story.effort or '',
                'priority': user_story.priority,
                'created_at': user_story.created_at.strftime('%Y-%m-%d %H:%M:%S') if user_story.created_at else ''
            }
            
            return jsonify({'success': True, 'message': '用户故事添加成功', 'user_story': story_data})
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'message': f'添加失败: {str(e)}'})
    
    # GET请求返回添加用户故事模态框
    return render_template('add_user_story_modal.html', 
                          users=users,
                          story_id=f"US_{product_backlog_id:02d}_")

@user_stories_bp.route('/edit_user_story/<int:user_story_id>', methods=['POST'])
def edit_user_story(user_story_id):
    # 检查权限
    if not check_system_feature_access(session, 'user_stories.user_stories'):
        return jsonify({'success': False, 'message': '权限不足'})
    
    user_story = db.session.get(UserStory, user_story_id)
    if user_story:
        title = request.form.get('title')
        story_id = request.form.get('story_id')
        description = request.form.get('description')
        acceptance_criteria = request.form.get('acceptance_criteria')
        effort = request.form.get('effort')
        priority = request.form.get('priority')
        iteration = request.form.get('iteration')
        
        if title:
            user_story.title = title
        if story_id is not None:
            user_story.story_id = story_id
        if description is not None:
            user_story.description = description
        if acceptance_criteria is not None:
            user_story.acceptance_criteria = acceptance_criteria
        if effort is not None:
            # 处理空字符串或None值
            if effort == '' or effort is None:
                user_story.effort = None
            else:
                try:
                    user_story.effort = float(effort)
                except (ValueError, TypeError):
                    user_story.effort = None
        if priority is not None:
            user_story.priority = priority
            
        db.session.commit()
        
        # 返回更新后的用户故事信息
        updated_story_data = {
            'id': user_story.id,
            'story_id': user_story.story_id or '',
            'title': user_story.title,
            'description': user_story.description or '',
            'priority': user_story.priority,
            'effort': user_story.effort,
            'created_at': user_story.created_at.strftime('%Y-%m-%d %H:%M:%S') if user_story.created_at else ''
        }
        
        return jsonify({'success': True, 'message': '用户故事更新成功', 'user_story': updated_story_data})
    else:
        return jsonify({'success': False, 'message': '用户故事不存在'})

@user_stories_bp.route('/delete_user_story/<int:user_story_id>', methods=['POST'])
def delete_user_story(user_story_id):
    # 检查权限
    if not check_system_feature_access(session, 'user_stories.user_stories'):
        return jsonify({'success': False, 'message': '权限不足'})
    
    user_story = db.session.get(UserStory, user_story_id)
    if user_story:
        # 检查用户故事是否已经加入某个迭代
        sprint_backlogs = SprintBacklog.query.filter_by(user_story_id=user_story_id).all()

        if sprint_backlogs:
            # 获取相关的迭代名称
            sprint_names = [sb.sprint.name for sb in sprint_backlogs]
            sprint_names_str = ', '.join(sprint_names)

            return jsonify({
                'success': False,
                'message': f'该故事已经加入了 {sprint_names_str} 迭代，请到迭代管理中去做相对应操作'
            })

        try:
            db.session.delete(user_story)
            db.session.commit()
            return jsonify({'success': True, 'message': '用户故事删除成功'})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'message': f'删除失败: {str(e)}'})
    else:
        return jsonify({'success': False, 'message': '用户故事不存在'})

@user_stories_bp.route('/get_edit_user_story_modal/<int:user_story_id>')
def get_edit_user_story_modal(user_story_id):
    """获取编辑用户故事的模态框"""
    # 检查权限
    if not check_system_feature_access(session, 'user_stories.user_stories'):
        return jsonify({'success': False, 'message': '权限不足'})
    
    # 获取用户故事
    user_story = db.session.get(UserStory, user_story_id)
    if not user_story:
        return jsonify({'success': False, 'message': '用户故事不存在'})
    
    # 获取所有非管理员用户，用于分配负责人
    users = User.query.join(UserRole).filter(Role.name != 'admin').all()
    
    return render_template('edit_user_story_modal.html', 
                         user_story=user_story, 
                         users=users)


@user_stories_bp.route('/get_add_user_story_modal')
def get_add_user_story_modal():
    """获取添加用户故事的模态框"""
    # 检查权限
    if not check_system_feature_access(session, 'user_stories.user_stories'):
        return jsonify({'success': False, 'message': '权限不足'})
    
    # 获取页面ID参数
    project_page_id = request.args.get('project_page_id', type=int)
    
    # 获取所有非管理员用户，用于分配负责人
    users = User.query.join(UserRole).filter(Role.name != 'admin').all()
    
    # 获取项目页面信息
    project_page = db.session.get(ProjectInfo, project_page_id) if project_page_id else None
    # 生成故事编号
    story_id = None
    if project_page_id:
        story_id = generate_story_id(project_page_id)
    return render_template('add_user_story_modal.html', 
                         users=users,
                         story_id=story_id )

@user_stories_bp.route('/export_user_stories/<int:page_id>')
def export_user_stories(page_id):
    """导出指定页面的用户故事到Excel文件"""
    # 检查权限
    if not check_system_feature_access(session, 'user_stories.user_stories'):
        return redirect(url_for('auth.index'))
    
    # 获取页面节点
    page_node = db.session.get(ProjectInfo, page_id)
    if not page_node or page_node.node_type != 'page':
        flash('页面不存在或不是有效的页面节点', 'error')
        return redirect(url_for('user_stories.user_stories'))
    
    # 获取页面关联的用户故事
    user_stories = UserStory.query.filter_by(project_page_id=page_id).order_by(UserStory.created_at.desc()).all()
    
    # 获取项目和菜单信息
    # 从路径中解析项目、菜单和页面信息
    path_parts = page_node.path.split('/')
    project_name = path_parts[1] if len(path_parts) > 1 else ''
    menu_name = path_parts[2] if len(path_parts) > 2 else ''
    page_name = page_node.name
    
    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "用户故事"
    
    # 设置标题行样式
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入标题行
    headers = ['项目', '菜单', '页面', 'ID', '故事编号', '故事标题', '故事描述', '验收标准', '优先级', '工作量', '创建时间']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 写入数据行
    for row, story in enumerate(user_stories, 2):
        ws.cell(row=row, column=1, value=project_name)
        ws.cell(row=row, column=2, value=menu_name)
        ws.cell(row=row, column=3, value=page_name)
        ws.cell(row=row, column=4, value=story.id)
        ws.cell(row=row, column=5, value=story.story_id or '')
        ws.cell(row=row, column=6, value=story.title or '')
        ws.cell(row=row, column=7, value=story.description or '')
        ws.cell(row=row, column=8, value=story.acceptance_criteria or '')
        ws.cell(row=row, column=9, value=story.priority or '')
        ws.cell(row=row, column=10, value=str(story.effort) if story.effort else '')
        ws.cell(row=row, column=11, value=story.created_at.strftime('%Y-%m-%d %H:%M:%S') if story.created_at else '')
    
    # 调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = min(adjusted_width, 50)
    
    # 将工作簿保存到内存中
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 生成文件名
    filename = f"用户故事_{project_name}_{menu_name}_{page_name}_{page_node.id}.xlsx"
    
    # 返回Excel文件作为附件下载
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
