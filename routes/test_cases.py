from flask import Blueprint, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from sqlalchemy import or_
from models import db, TestCase, User, ProjectInfo, Sprint, UserStory, SprintBacklog, ProductBacklog
from utils import check_system_feature_access
from decorators import check_access_blueprint
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO

test_cases_bp = Blueprint('test_cases', __name__)

# 应用权限检查装饰器
@test_cases_bp.before_request
@check_access_blueprint('test_cases')
def check_access():
    pass  # 装饰器会处理权限检查逻辑

def generate_case_id(project_short_name, user_story_id):
    """
    生成测试用例编号
    规则：项目简称-用户故事ID-001，如：TIM-001-001
    """
    if not project_short_name or not user_story_id:
        return None

    # 查询当前用户故事下最大的测试用例序号
    cases = TestCase.query.filter_by(user_story_id=user_story_id).all()

    if cases:
        # 查找最大的序号并加1
        max_seq = 0
        for case in cases:
            if case.case_id and case.case_id.startswith(f"{project_short_name}-{user_story_id}-"):
                try:
                    seq = int(case.case_id.split('-')[2])
                    if seq > max_seq:
                        max_seq = seq
                except (IndexError, ValueError):
                    # 如果解析失败，跳过该用例
                    pass
        next_seq = max_seq + 1
    else:
        # 如果当前用户故事下没有测试用例，默认值为1
        next_seq = 1

    # 序号保持3位长度，不够前面补零
    seq_part = str(next_seq).zfill(3)

    # 拼接生成用例编号
    case_id = f"{project_short_name}-{user_story_id}-{seq_part}"
    return case_id

@test_cases_bp.route('/test_cases')
def test_cases():
    """测试用例列表页面"""
    # 检查权限
    if not check_system_feature_access(session, 'test_cases.test_cases'):
        return redirect(url_for('auth.index'))

    # 获取查询参数
    page = request.args.get('page', 1, type=int)
    per_page = 10
    search = request.args.get('search', '', type=str)
    project_id = request.args.get('project_id', 0, type=int)
    sprint_id = request.args.get('sprint_id', 0, type=int)

    # 构建查询
    query = TestCase.query

    # 搜索条件
    if search:
        query = query.filter(
            or_(
                TestCase.case_id.contains(search),
                TestCase.title.contains(search)
            )
        )

    # 项目筛选
    if project_id:
        query = query.filter(TestCase.project_id == project_id)

    # 迭代筛选
    if sprint_id:
        query = query.filter(TestCase.sprint_id == sprint_id)

    # 分页查询
    pagination = query.order_by(TestCase.case_id).paginate(
        page=page, per_page=per_page, error_out=False)
    test_cases = pagination.items

    # 获取项目和迭代列表用于筛选
    projects = ProjectInfo.query.filter_by(parent_id=None).all()
    sprints = Sprint.query.all()

    return render_template('test_cases.html',
                          test_cases=test_cases,
                          pagination=pagination,
                          projects=projects,
                          sprints=sprints,
                          search=search,
                          project_id=project_id,
                          sprint_id=sprint_id)

@test_cases_bp.route('/test_cases/add', methods=['GET', 'POST'])
def add_test_case():
    """添加测试用例"""
    # 检查权限
    if not check_system_feature_access(session, 'test_cases.test_cases'):
        return redirect(url_for('auth.index'))

    if request.method == 'POST':
        # 获取表单数据
        project_id = request.form.get('project_id', type=int)
        sprint_id = request.form.get('sprint_id', type=int)
        user_story_id = request.form.get('user_story_id', type=int)
        case_type = request.form.get('case_type')
        function_point = request.form.get('function_point')
        title = request.form.get('title')
        precondition = request.form.get('precondition')
        steps = request.form.get('steps')
        expected_result = request.form.get('expected_result')
        priority = request.form.get('priority', 'P3')
        is_automated = request.form.get('is_automated') == 'on'
        remarks = request.form.get('remarks')

        # 获取项目简称用于生成用例编号
        project_short_name = None
        if project_id:
            project = ProjectInfo.query.get(project_id)
            if project and project.short_name:
                project_short_name = project.short_name

        # 生成用例编号
        case_id = None
        user_story = None
        if user_story_id:
            user_story = UserStory.query.get(user_story_id)
            if user_story and user_story.story_id and project_short_name:
                case_id = generate_case_id(project_short_name, user_story.story_id)

        # 创建测试用例对象
        test_case = TestCase(
            case_id=case_id,
            project_id=project_id,
            sprint_id=sprint_id if sprint_id else None,
            user_story_id=user_story_id if user_story_id else None,
            case_type=case_type,
            function_point=function_point,
            title=title,
            precondition=precondition,
            steps=steps,
            expected_result=expected_result,
            priority=priority,
            is_automated=is_automated,
            remarks=remarks,
            created_by_id=session.get('user_id')
        )

        try:
            db.session.add(test_case)
            db.session.commit()
            flash('测试用例添加成功！', 'success')
            return redirect(url_for('test_cases.test_cases'))
        except Exception as e:
            db.session.rollback()
            flash('添加测试用例失败，请重试！', 'error')

    # GET请求，显示添加页面
    projects = ProjectInfo.query.filter_by(parent_id=None).all()
    sprints = Sprint.query.all()
    user_stories = UserStory.query.all()
    users = User.query.all()

    return render_template('test_case_form.html',
                          action='add',
                          projects=projects,
                          sprints=sprints,
                          user_stories=user_stories,
                          users=users)

@test_cases_bp.route('/test_cases/edit/<int:case_id>', methods=['GET', 'POST'])
def edit_test_case(case_id):
    """编辑测试用例"""
    # 检查权限
    if not check_system_feature_access(session, 'test_cases.test_cases'):
        return redirect(url_for('auth.index'))

    # 获取要编辑的测试用例
    test_case = TestCase.query.get_or_404(case_id)

    if request.method == 'POST':
        # 获取表单数据
        project_id = request.form.get('project_id', type=int)
        sprint_id = request.form.get('sprint_id', type=int)
        user_story_id = request.form.get('user_story_id', type=int)
        case_type = request.form.get('case_type')
        function_point = request.form.get('function_point')
        title = request.form.get('title')
        precondition = request.form.get('precondition')
        steps = request.form.get('steps')
        expected_result = request.form.get('expected_result')
        actual_result = request.form.get('actual_result')
        test_environment = request.form.get('test_environment')
        priority = request.form.get('priority')
        is_automated = request.form.get('is_automated') == 'on'
        edit_status = request.form.get('edit_status')
        execution_status = request.form.get('execution_status')
        test_result = request.form.get('test_result')
        tested_by_id = request.form.get('tested_by_id', type=int)
        tested_at = request.form.get('tested_at')
        remarks = request.form.get('remarks')

        # 更新测试用例对象
        test_case.project_id = project_id
        test_case.sprint_id = sprint_id if sprint_id else None
        test_case.user_story_id = user_story_id if user_story_id else None
        test_case.case_type = case_type
        test_case.function_point = function_point
        test_case.title = title
        test_case.precondition = precondition
        test_case.steps = steps
        test_case.expected_result = expected_result
        test_case.actual_result = actual_result
        test_case.test_environment = test_environment
        test_case.priority = priority
        test_case.is_automated = is_automated
        test_case.edit_status = edit_status
        test_case.execution_status = execution_status
        test_case.test_result = test_result
        test_case.tested_by_id = tested_by_id if tested_by_id else None

        # 处理测试时间
        if tested_at:
            try:
                test_case.tested_at = datetime.strptime(tested_at, '%Y-%m-%d')
            except ValueError:
                test_case.tested_at = None
        else:
            test_case.tested_at = None

        test_case.remarks = remarks

        try:
            db.session.commit()
            flash('测试用例更新成功！', 'success')
            return redirect(url_for('test_cases.test_cases'))
        except Exception as e:
            db.session.rollback()
            flash('更新测试用例失败，请重试！', 'error')

    # GET请求，显示编辑页面
    projects = ProjectInfo.query.filter_by(parent_id=None).all()
    sprints = Sprint.query.all()
    user_stories = UserStory.query.all()
    users = User.query.all()

    return render_template('test_case_form.html',
                          action='edit',
                          test_case=test_case,
                          projects=projects,
                          sprints=sprints,
                          user_stories=user_stories,
                          users=users)

@test_cases_bp.route('/test_cases/delete/<int:case_id>', methods=['POST'])
def delete_test_case(case_id):
    """删除测试用例"""
    # 检查权限
    if not check_system_feature_access(session, 'test_cases.test_cases'):
        return jsonify({'success': False, 'message': '权限不足'})

    # 获取要删除的测试用例
    test_case = TestCase.query.get_or_404(case_id)

    try:
        db.session.delete(test_case)
        db.session.commit()
        return jsonify({'success': True, 'message': '测试用例删除成功！'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': '删除测试用例失败，请重试！'})

@test_cases_bp.route('/test_cases/get_sprints')
def get_sprints():
    """根据项目ID获取迭代"""
    # 检查权限
    if not check_system_feature_access(session, 'test_cases.test_cases'):
        return jsonify({'success': False, 'message': '权限不足'})

    project_id = request.args.get('project_id', 0, type=int)

    if not project_id:
        return jsonify({'success': False, 'message': '项目ID不能为空'})

    # 获取项目下的所有迭代
    sprints = Sprint.query.filter_by(project_id=project_id).all()

    sprints_data = []
    for sprint in sprints:
        sprints_data.append({
            'id': sprint.id,
            'name': sprint.name
        })

    return jsonify({'success': True, 'sprints': sprints_data})


@test_cases_bp.route('/test_cases/get_sprints_by_project')
def get_sprints_by_project():
    """根据项目ID获取迭代"""
    # 检查权限
    if not check_system_feature_access(session, 'test_cases.test_cases'):
        return jsonify({'success': False, 'message': '权限不足'})

    project_id = request.args.get('project_id', 0, type=int)

    print(f"获取迭代，项目ID: {project_id}")  # 调试信息

    if not project_id:
        return jsonify({'success': False, 'message': '项目ID不能为空'})

    try:
        # 获取指定项目下的迭代
        sprints = Sprint.query.filter_by(project_id=project_id).all()

        print(f"找到 {len(sprints)} 个迭代")  # 调试信息

        sprints_data = []
        for sprint in sprints:
            sprints_data.append({
                'id': sprint.id,
                'name': sprint.name
            })

        return jsonify({'success': True, 'sprints': sprints_data})
    except Exception as e:
        print(f"获取迭代时出错: {str(e)}")  # 调试信息
        return jsonify({'success': False, 'message': '服务器内部错误'})


@test_cases_bp.route('/test_cases/get_user_stories_by_sprint')
def get_user_stories_by_sprint():
    """根据迭代ID获取用户故事"""
    # 检查权限
    if not check_system_feature_access(session, 'test_cases.test_cases'):
        return jsonify({'success': False, 'message': '权限不足'})

    sprint_id = request.args.get('sprint_id', 0, type=int)

    print(f"获取用户故事，迭代ID: {sprint_id}")  # 调试信息

    if not sprint_id:
        return jsonify({'success': False, 'message': '迭代ID不能为空'})

    try:
        # 获取迭代下的所有用户故事
        # 通过SprintBacklog关联Sprint和UserStory
        sprint_backlogs = SprintBacklog.query.filter_by(sprint_id=sprint_id).all()
        user_stories = [sb.user_story for sb in sprint_backlogs if sb.user_story]

        print(f"找到 {len(user_stories)} 个用户故事")  # 调试信息

        stories_data = []
        for story in user_stories:
            stories_data.append({
                'id': story.id,
                'story_id': story.story_id or '',
                'title': story.title or ''
            })

        return jsonify({'success': True, 'user_stories': stories_data})
    except Exception as e:
        print(f"获取用户故事时出错: {str(e)}")  # 调试信息
        return jsonify({'success': False, 'message': '服务器内部错误'})



@test_cases_bp.route('/test_cases/get_user_stories')
def get_user_stories():
    """根据项目ID获取用户故事"""
    # 检查权限
    if not check_system_feature_access(session, 'test_cases.test_cases'):
        return jsonify({'success': False, 'message': '权限不足'})

    project_id = request.args.get('project_id', 0, type=int)

    if not project_id:
        return jsonify({'success': False, 'message': '项目ID不能为空'})

    # 获取项目下的所有页面
    pages = ProjectInfo.query.filter_by(parent_id=project_id).all()
    page_ids = [page.id for page in pages]

    # 获取页面下的所有用户故事
    user_stories = UserStory.query.join(ProductBacklog).filter(
        ProductBacklog.project_module_id.in_(page_ids)
    ).all()

    stories_data = []
    for story in user_stories:
        stories_data.append({
            'id': story.id,
            'story_id': story.story_id,
            'title': story.title
        })

    return jsonify({'success': True, 'user_stories': stories_data})

@test_cases_bp.route('/test_cases/export')
def export_test_cases():
    """导出测试用例到Excel文件"""
    # 检查权限
    if not check_system_feature_access(session, 'test_cases.test_cases'):
        return redirect(url_for('auth.index'))

    # 获取查询参数
    project_id = request.args.get('project_id', 0, type=int)
    sprint_id = request.args.get('sprint_id', 0, type=int)

    # 构建查询
    query = TestCase.query

    # 项目筛选
    if project_id:
        query = query.filter(TestCase.project_id == project_id)

    # 迭代筛选
    if sprint_id:
        query = query.filter(TestCase.sprint_id == sprint_id)

    # 获取测试用例数据
    test_cases = query.order_by(TestCase.created_at.desc()).all()

    # 获取项目和迭代信息
    project_name = ""
    sprint_name = ""

    if project_id:
        project = ProjectInfo.query.get(project_id)
        if project:
            project_name = project.name

    if sprint_id:
        sprint = Sprint.query.get(sprint_id)
        if sprint:
            sprint_name = sprint.name

    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # 设置标题行样式
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 写入标题行
    headers = [
        '用例编号', '用例编辑状态', '所属项目', '项目模块', '所属迭代', '用户故事',
        '用例标题', '测试用例类型', '具体功能点', '预置条件', '测试步骤', '预期结果',
        '优先级', '是否自动化（是/否）', '编写人', '编写时间', '测试人', '测试环境',
        '测试时间', '用例执行状态', '测试结果','实际结果', '备注'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = header_alignment

        # 写入数据行
        for row, case in enumerate(test_cases, 2):
            # 获取关联信息
            project_name = case.project.name if case.project else ''
            sprint_name = case.sprint.name if case.sprint else ''
            user_story_title = case.user_story.title if case.user_story else ''
            created_by_name = case.created_by.name if case.created_by else ''
            tested_by_name = case.tested_by.name if case.tested_by else ''

            ws.cell(row=row, column=1, value=case.case_id or '')
            ws.cell(row=row, column=2, value=case.edit_status or '')
            ws.cell(row=row, column=3, value=project_name)
            ws.cell(row=row, column=4, value=case.project_module or '')
            ws.cell(row=row, column=5, value=sprint_name)
            ws.cell(row=row, column=6, value=user_story_title)
            ws.cell(row=row, column=7, value=case.title or '')
            ws.cell(row=row, column=8, value=case.case_type or '')
            ws.cell(row=row, column=9, value=case.function_point or '')
            ws.cell(row=row, column=10, value=case.precondition or '')
            ws.cell(row=row, column=11, value=case.steps or '')
            ws.cell(row=row, column=12, value=case.expected_result or '')
            ws.cell(row=row, column=13, value=case.priority or '')
            ws.cell(row=row, column=14, value='是' if case.is_automated else '否')
            ws.cell(row=row, column=15, value=created_by_name)
            ws.cell(row=row, column=16, value=case.created_at.strftime('%Y-%m-%d %H:%M:%S') if case.created_at else '')
            ws.cell(row=row, column=17, value=tested_by_name)
            ws.cell(row=row, column=18, value=case.test_environment or '')
            ws.cell(row=row, column=19, value=case.tested_at.strftime('%Y-%m-%d') if case.tested_at else '')
            ws.cell(row=row, column=20, value=case.execution_status or '')
            ws.cell(row=row, column=21, value=case.test_result or '')
            ws.cell(row=row, column=22, value=case.actual_result or '')
            ws.cell(row=row, column=23, value=case.remarks or '')

    # 调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = min(adjusted_width, 50)

    # 将工作簿保存到内存中
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # 生成文件名
    filename_parts = ["测试用例"]
    if project_name:
        filename_parts.append(project_name)
    if sprint_name:
        filename_parts.append(sprint_name)
    filename = "_".join(filename_parts) + ".xlsx"

    # 返回Excel文件作为附件下载
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@test_cases_bp.route('/test_cases/import', methods=['GET', 'POST'])
def import_test_cases():
    """从Excel文件导入测试用例"""
    # 检查权限
    if not check_system_feature_access(session, 'test_cases.test_cases'):
        return redirect(url_for('auth.index'))

    if request.method == 'POST':
        # 检查是否有文件上传
        if 'file' not in request.files:
            flash('请选择要上传的文件', 'error')
            return redirect(request.url)

        file = request.files['file']

        # 检查文件名
        if file.filename == '':
            flash('请选择要上传的文件', 'error')
            return redirect(request.url)

        # 检查文件类型
        if not file.filename.endswith('.xlsx'):
            flash('只支持.xlsx格式的文件', 'error')
            return redirect(request.url)

        try:
            # 读取Excel文件
            workbook = load_workbook(file)
            worksheet = workbook.active

            # 解析数据并创建测试用例
            success_count = 0
            error_count = 0

            # 用于跟踪当前导入过程中每个用户故事的用例序号
            import_case_counters = {}

            # 从第二行开始读取数据（第一行是标题）
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 2):
                try:
                    # 获取行数据
                    case_id = row[0] if row[0] else None
                    edit_status = row[1] if row[1] else '新增'
                    project_name = row[2] if row[2] else None
                    project_module = row[3] if row[3] else None
                    sprint_name = row[4] if row[4] else None
                    user_story_title = row[5] if row[5] else None
                    title = row[6] if row[6] else None
                    case_type = row[7] if row[7] else None
                    function_point = row[8] if row[8] else None
                    precondition = row[9] if row[9] else None
                    steps = row[10] if row[10] else None
                    expected_result = row[11] if row[11] else None
                    priority = row[12] if row[12] else 'P3'
                    is_automated_str = row[13] if row[13] else '否'
                    created_by_name = row[14] if row[14] else None
                    created_at_str = row[15] if row[15] else None
                    tested_by_name = row[16] if row[16] else None
                    test_environment = row[17] if row[17] else None
                    tested_at_str = row[18] if row[18] else None
                    execution_status = row[19] if row[19] else '未开始'
                    test_result = row[20] if row[20] else None
                    actual_result = row[21] if row[21] else None
                    remarks = row[22] if row[22] else None

                    # 验证必填字段
                    if not title:
                        error_count += 1
                        continue

                    # 查找关联对象
                    project = None
                    project_short_name = None
                    if project_name:
                        project = ProjectInfo.query.filter_by(name=project_name).first()
                        if project and project.short_name:
                            project_short_name = project.short_name

                    sprint = None
                    if sprint_name:
                        sprint = Sprint.query.filter_by(name=sprint_name).first()

                    user_story = None
                    user_story_id = None
                    if user_story_title:
                        user_story = UserStory.query.filter_by(title=user_story_title).first()
                        if user_story and user_story.story_id:
                            user_story_id = user_story.story_id

                    created_by = None
                    if created_by_name:
                        created_by = User.query.filter_by(name=created_by_name).first()

                    tested_by = None
                    if tested_by_name:
                        tested_by = User.query.filter_by(name=tested_by_name).first()

                    # 处理布尔值
                    is_automated = is_automated_str == '是'

                    # 处理日期
                    created_at = None
                    if created_at_str:
                        try:
                            created_at = datetime.strptime(created_at_str, '%Y-%m-%d %H:%M:%S')
                        except ValueError:
                            pass

                    tested_at = None
                    if tested_at_str:
                        try:
                            tested_at = datetime.strptime(tested_at_str, '%Y-%m-%d')
                        except ValueError:
                            pass

                    # 自动生成用例编号
                    generated_case_id = None
                    if project_short_name and user_story_id:
                        # 创建一个用于当前导入过程的唯一键
                        counter_key = f"{project_short_name}-{user_story_id}"

                        # 检查是否已经为这个用户故事生成过用例编号
                        if counter_key in import_case_counters:
                            # 如果已经存在，递增计数器
                            import_case_counters[counter_key] += 1
                        else:
                            # 如果不存在，先查询数据库中已有的最大序号
                            max_seq = 0
                            existing_cases = TestCase.query.filter_by(
                                user_story_id=user_story.id if user_story else None).all()
                            for case in existing_cases:
                                if case.case_id and case.case_id.startswith(f"{project_short_name}-{user_story_id}-"):
                                    try:
                                        seq = int(case.case_id.split('-')[2])
                                        if seq > max_seq:
                                            max_seq = seq
                                    except (IndexError, ValueError):
                                        # 如果解析失败，跳过该用例
                                        pass
                            # 设置初始计数器为数据库中最大序号+1
                            import_case_counters[counter_key] = max_seq + 1

                        # 使用当前计数器值生成用例编号
                        next_seq = import_case_counters[counter_key]
                        seq_part = str(next_seq).zfill(3)
                        generated_case_id = f"{project_short_name}-{user_story_id}-{seq_part}"
                    elif case_id:
                        # 如果无法自动生成，则使用Excel中提供的用例编号
                        generated_case_id = case_id
                    # print(generated_case_id)

                    # 创建测试用例对象
                    test_case = TestCase(
                        case_id=generated_case_id,  # 使用自动生成的用例编号
                        project_id=project.id if project else None,
                        project_module=project_module,
                        sprint_id=sprint.id if sprint else None,
                        user_story_id=user_story.id if user_story else None,
                        edit_status=edit_status,
                        execution_status=execution_status,
                        test_result=test_result,
                        case_type=case_type,
                        function_point=function_point,
                        title=title,
                        precondition=precondition,
                        steps=steps,
                        expected_result=expected_result,
                        actual_result=actual_result,
                        test_environment=test_environment,
                        priority=priority,
                        is_automated=is_automated,
                        created_by_id=created_by.id if created_by else session.get('user_id'),
                        created_at=created_at,
                        tested_by_id=tested_by.id if tested_by else None,
                        tested_at=tested_at,
                        remarks=remarks
                    )

                    db.session.add(test_case)
                    success_count += 1

                except Exception as e:
                    error_count += 1
                    error_details = {
                        'row_number': row_num,
                        'error_message': str(e),
                        'row_data': row  # 保存出错行的数据以便分析
                    }
                    # 记录错误但继续处理其他行
                    print(f"处理第{row_num}行时出错: {error_details}")

            db.session.commit()
            flash(f'导入完成！成功导入{success_count}条记录，{error_count}条记录失败。', 'success')

        except Exception as e:
            db.session.rollback()
            flash(f'导入失败：{str(e)}', 'error')

        return redirect(url_for('test_cases.test_cases'))

    # GET请求显示导入页面
    return render_template('test_case_import.html')
