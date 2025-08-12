from flask import Blueprint, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from openpyxl.reader.excel import load_workbook

from models import db, ProductBacklog, User, ProjectInfo
from utils import check_system_feature_access, check_user_role
from decorators import check_access_blueprint
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO
import json

product_backlog_bp = Blueprint('product_backlog', __name__)

# 应用权限检查装饰器
@product_backlog_bp.before_request
@check_access_blueprint('product_backlog')
def check_access():
    pass  # 装饰器会处理权限检查逻辑

@product_backlog_bp.route('/product_backlog')
def product_backlog():
    # 检查权限
    if not check_system_feature_access(session, 'product_backlog.product_backlog'):
        return redirect(url_for('auth.index'))

    # 获取所有项目（根节点）
    projects = ProjectInfo.query.filter_by(parent_id=None).order_by(ProjectInfo.order).all()

    # 获取所有产品待办事项
    product_backlogs = ProductBacklog.query.order_by(ProductBacklog.created_at.desc()).all()

    # 获取所有非管理员用户，用于分配责任人和分析人员
    all_users = User.query.all()
    users = []
    for user in all_users:
        if not check_user_role(user.id, 'admin'):
            users.append(user)

    return render_template('product_backlog.html',
                          product_backlogs=product_backlogs,
                          users=users,
                          projects=projects)


def generate_requirement_id():
    """
    生成产品待办事项编号
    规则：R_序号(3位)
    """
    # 查询当前最大的需求序号
    backlogs = ProductBacklog.query.all()

    if backlogs:
        # 查找最大的序号并加1
        max_seq = 0
        for backlog in backlogs:
            if backlog.requirement_id and backlog.requirement_id.startswith("R_"):
                try:
                    seq = int(backlog.requirement_id.split('_')[1])
                    if seq > max_seq:
                        max_seq = seq
                except (IndexError, ValueError):
                    # 如果解析失败，跳过该需求
                    pass
        next_seq = max_seq + 1
    else:
        # 如果没有产品待办事项，默认值为1
        next_seq = 1

    # 序号保持3位长度，不够前面补零
    seq_part = str(next_seq).zfill(3)

    # 拼接生成需求编号
    requirement_id = f"R_{seq_part}"
    return requirement_id

@product_backlog_bp.route('/product_backlog/add', methods=['POST'])
def add_product_backlog():
    # 检查权限
    if not check_system_feature_access(session, 'product_backlog.product_backlog'):
        return jsonify({'success': False, 'message': '权限不足'})

    # 获取表单数据
    title = request.form.get('title')
    description = request.form.get('description')
    requirement_type = request.form.get('requirement_type')
    customer_owner_id = request.form.get('customer_owner_id')
    priority = request.form.get('priority', 'P3')
    status = request.form.get('status', '待讨论')
    project_id = request.form.get('project_id')
    project_module_id = request.form.get('project_module_id')
    analyst_id = request.form.get('analyst_id')
    progress = request.form.get('progress', '未处理')
    related_info = request.form.get('related_info')
    tags = request.form.get('tags')

    # 验证必要字段
    if not title:
        return jsonify({'success': False, 'message': '需求标题不能为空'})

    # 生成需求编号
    requirement_id = generate_requirement_id()

    # 创建新的产品待办事项
    backlog = ProductBacklog(
        requirement_id=requirement_id,
        title=title,
        description=description,
        requirement_type=requirement_type,
        customer_owner_id=int(customer_owner_id) if customer_owner_id else None,
        priority=priority,
        status=status,
        project_id=int(project_id) if project_id else None,
        project_module_id=int(project_module_id) if project_module_id else None,
        analyst_id=int(analyst_id) if analyst_id else None,
        progress=progress,
        related_info=related_info,
        tags=tags
    )

    try:
        db.session.add(backlog)
        db.session.commit()
        return jsonify({'success': True, 'message': '需求添加成功', 'requirement_id': requirement_id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'添加失败: {str(e)}'})

@product_backlog_bp.route('/product_backlog/get/<int:backlog_id>')
def get_product_backlog(backlog_id):
    try:
        # 检查权限
        if not check_system_feature_access(session, 'product_backlog.product_backlog'):
            return jsonify({'success': False, 'message': '权限不足'})

        # 获取产品待办事项
        backlog = db.session.get(ProductBacklog, backlog_id)
        if not backlog:
            return jsonify({'success': False, 'message': '需求不存在'})

        # 构造返回数据
        backlog_data = {
            'id': backlog.id,
            'requirement_id': backlog.requirement_id,
            'title': backlog.title,
            'description': backlog.description,
            'requirement_type': backlog.requirement_type,
            'customer_owner_id': backlog.customer_owner_id,
            'priority': backlog.priority,
            'status': backlog.status,
            'project_id': backlog.project_id,
            'project_module_id': backlog.project_module_id,
            'analyst_id': backlog.analyst_id,
            'progress': backlog.progress,
            'related_info': backlog.related_info,
            'tags': backlog.tags
        }

        return jsonify({'success': True, 'backlog': backlog_data})
    except Exception as e:
        print(f"Error in get_product_backlog: {str(e)}")
        return jsonify({'success': False, 'message': f'获取需求信息失败: {str(e)}'}), 500


@product_backlog_bp.route('/product_backlog/edit/<int:backlog_id>', methods=['POST'])
def edit_product_backlog(backlog_id):
    # 检查权限
    if not check_system_feature_access(session, 'product_backlog.product_backlog'):
        return jsonify({'success': False, 'message': '权限不足'})

    # 获取产品待办事项
    backlog = db.session.get(ProductBacklog, backlog_id)
    if not backlog:
        return jsonify({'success': False, 'message': '需求不存在'})

    # 获取表单数据
    title = request.form.get('title')
    description = request.form.get('description')
    requirement_type = request.form.get('requirement_type')
    customer_owner_id = request.form.get('customer_owner_id')
    priority = request.form.get('priority')
    status = request.form.get('status')
    project_id = request.form.get('project_id')
    project_module_id = request.form.get('project_module_id')
    analyst_id = request.form.get('analyst_id')
    progress = request.form.get('progress')
    related_info = request.form.get('related_info')
    tags = request.form.get('tags')

    # 验证必要字段
    if not title:
        return jsonify({'success': False, 'message': '需求标题不能为空'})

    # 更新产品待办事项
    backlog.title = title
    backlog.description = description
    backlog.requirement_type = requirement_type
    backlog.customer_owner_id = int(customer_owner_id) if customer_owner_id else None
    backlog.priority = priority
    backlog.status = status
    backlog.project_id = int(project_id) if project_id else None
    backlog.project_module_id = int(project_module_id) if project_module_id else None
    backlog.analyst_id = int(analyst_id) if analyst_id else None
    backlog.progress = progress
    backlog.related_info = related_info
    backlog.tags = tags

    try:
        db.session.commit()
        return jsonify({'success': True, 'message': '需求更新成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'更新失败: {str(e)}'})

@product_backlog_bp.route('/product_backlog/delete/<int:backlog_id>', methods=['POST'])
def delete_product_backlog(backlog_id):
    # 检查权限
    if not check_system_feature_access(session, 'product_backlog.product_backlog'):
        return jsonify({'success': False, 'message': '权限不足'})

    # 获取产品待办事项
    backlog = db.session.get(ProductBacklog, backlog_id)
    if not backlog:
        return jsonify({'success': False, 'message': '需求不存在'})

    try:
        db.session.delete(backlog)
        db.session.commit()
        return jsonify({'success': True, 'message': '需求删除成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'删除失败: {str(e)}'})

@product_backlog_bp.route('/product_backlog/export')
def export_product_backlog():
    # 检查权限
    if not check_system_feature_access(session, 'product_backlog.product_backlog'):
        return redirect(url_for('auth.index'))

    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "产品待办事项"

    # 设置表头样式
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 写入表头
    headers = ['需求编号', '所属项目', '功能模块', '需求标题', '需求描述', '需求类型', '责任人', '优先级',
              '提出日期', '需求状态', '分析人员', '执行进度', '关联信息', '标签']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = header_alignment

    # 获取所有产品待办事项
    backlogs = ProductBacklog.query.order_by(ProductBacklog.created_at.desc()).all()

    # 写入数据
    for row, backlog in enumerate(backlogs, 2):
        ws.cell(row=row, column=1, value=backlog.requirement_id)
        ws.cell(row=row, column=2, value=backlog.project.name if backlog.project else '')
        ws.cell(row=row, column=3, value=backlog.project_module.name if backlog.project_module else '')
        ws.cell(row=row, column=4, value=backlog.title)
        ws.cell(row=row, column=5, value=backlog.description)
        ws.cell(row=row, column=6, value=backlog.requirement_type)
        ws.cell(row=row, column=7, value=backlog.customer_owner.name if backlog.customer_owner else '')
        ws.cell(row=row, column=8, value=backlog.priority)
        ws.cell(row=row, column=9, value=backlog.created_at.strftime('%Y-%m-%d') if backlog.created_at else '')
        ws.cell(row=row, column=10, value=backlog.status)
        ws.cell(row=row, column=11, value=backlog.analyst.name if backlog.analyst else '')
        ws.cell(row=row, column=12, value=backlog.progress)
        ws.cell(row=row, column=13, value=backlog.related_info)
        ws.cell(row=row, column=14, value=backlog.tags)

    # 设置列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = min(adjusted_width, 50)

    # 将工作簿保存到内存中
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # 返回Excel文件作为附件下载
    return send_file(output,
                    as_attachment=True,
                    download_name='产品待办事项.xlsx',
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@product_backlog_bp.route('/product_backlog/import', methods=['POST'])
def import_product_backlog():
    # 检查权限
    if not check_system_feature_access(session, 'product_backlog.product_backlog'):
        return jsonify({'success': False, 'message': '权限不足'})

    # 检查是否有上传文件
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '未选择文件'})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '未选择文件'})

    # 检查文件类型
    if not file.filename.endswith('.xlsx'):
        return jsonify({'success': False, 'message': '只支持Excel文件(.xlsx)'})

    try:
        # 读取Excel文件
        wb = load_workbook(file)
        ws = wb.active

        # 解析数据并导入到数据库
        imported_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            # 检查是否是空行
            if not any(cell for cell in row[:3]):
                continue

            # 创建新的产品待办事项
            backlog = ProductBacklog(
                requirement_id=row[0],
                title=row[1] or '',
                description=row[2] or '',
                requirement_type=row[3] or '',
                priority=row[5] or 'P3',
                status=row[7] or '待讨论',
                progress=row[10] or '未处理',
                related_info=row[11] or '',
                tags=row[12] or ''
            )

            # 处理责任人
            if row[4]:
                customer_owner = User.query.filter_by(name=str(row[4]).strip()).first()
                if customer_owner:
                    backlog.customer_owner_id = customer_owner.id

            # 处理所属项目
            if row[8]:
                project = ProjectInfo.query.filter_by(name=str(row[8]).strip()).first()
                if project:
                    backlog.project_id = project.id

            # 处理分析人员
            if row[9]:
                analyst = User.query.filter_by(name=str(row[9]).strip()).first()
                if analyst:
                    backlog.analyst_id = analyst.id

            db.session.add(backlog)
            imported_count += 1

        db.session.commit()
        return jsonify({'success': True, 'message': f'成功导入{imported_count}条需求'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'导入失败: {str(e)}'})
